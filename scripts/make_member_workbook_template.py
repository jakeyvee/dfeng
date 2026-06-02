#!/usr/bin/env python3
"""Generate the member-workbook Excel template from the canonical schema.

Produces ``config/member-workbook-template.xlsx`` with a ``Members`` tab whose
header row matches ``dfeng_bot.services.schema.MEMBER_COLUMNS`` EXACTLY (the bot
rejects header drift on startup), plus a ``Reference`` tab documenting the enums,
column ownership, and how to upload it to Google Drive as a Google Sheet.

Run:  python scripts/make_member_workbook_template.py
Deps: openpyxl  (pip install openpyxl)
"""
from __future__ import annotations

import sys
from pathlib import Path

# Make dfeng_bot importable whether or not the package is pip-installed.
ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT / "src"))

from dfeng_bot.services.schema import (  # noqa: E402
    ADMIN_COLUMNS,
    BOT_COLUMNS,
    ENTRY_SOURCES,
    KEY_COLUMN,
    MEMBER_COLUMNS,
    PII_COLUMNS,
    TAGS,
)
from openpyxl import Workbook  # noqa: E402
from openpyxl.comments import Comment  # noqa: E402
from openpyxl.styles import Alignment, Font, PatternFill  # noqa: E402
from openpyxl.utils import get_column_letter  # noqa: E402
from openpyxl.worksheet.datavalidation import DataValidation  # noqa: E402

OUT = ROOT / "config" / "member-workbook-template.xlsx"
TAB = "Members"
LAST_ROW = 1000  # rows the dropdowns + striping cover

# Brand palette: Dark Blue / Red / White (docs/branding-assets.md)
BOT_FILL = PatternFill("solid", fgColor="1F3864")    # dark blue  -> bot-owned
ADMIN_FILL = PatternFill("solid", fgColor="7C2128")  # deep red   -> admin-owned
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
WIDTHS = {
    "Telegram ID": 16, "Telegram username": 18, "Tag": 14, "Optional phone": 16,
    "Optional plate": 14, "Consent timestamp": 22, "Entry source": 18,
    "Joined timestamp": 22, "Notes": 28, "Status": 16, "Deletion requested": 18,
    "Last reconciled": 22,
}


def build() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = TAB

    # --- header row (must equal MEMBER_COLUMNS exactly) ---------------------
    for idx, name in enumerate(MEMBER_COLUMNS, start=1):
        cell = ws.cell(row=1, column=idx, value=name)
        cell.font = HEADER_FONT
        cell.fill = BOT_FILL if name in BOT_COLUMNS else ADMIN_FILL
        cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.column_dimensions[get_column_letter(idx)].width = WIDTHS.get(name, 16)
        note = None
        if name == KEY_COLUMN:
            note = "Unique key (bot-owned). Do not edit."
        elif name in PII_COLUMNS:
            note = "PII — keep private (PDPA). Optional; may be empty."
        elif name in BOT_COLUMNS:
            note = "Bot-owned: written by the bot. Treat as read-only."
        elif name in ADMIN_COLUMNS:
            note = "Admin-owned: edited by humans only. The bot never writes here."
        if note:
            cell.comment = Comment(note, "Dongfeng Bot")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(MEMBER_COLUMNS))}1"
    ws.sheet_view.showGridLines = True

    # --- dropdowns for the two enum columns --------------------------------
    def col_letter(header: str) -> str:
        return get_column_letter(MEMBER_COLUMNS.index(header) + 1)

    dv_tag = DataValidation(type="list", formula1='"%s"' % ",".join(TAGS), allow_blank=True)
    dv_tag.error = "Pick one of: " + ", ".join(TAGS)
    dv_tag.prompt = "Member tag"
    ws.add_data_validation(dv_tag)
    dv_tag.add(f"{col_letter('Tag')}2:{col_letter('Tag')}{LAST_ROW}")

    dv_src = DataValidation(type="list", formula1='"%s"' % ",".join(ENTRY_SOURCES), allow_blank=True)
    dv_src.error = "Pick one of: " + ", ".join(ENTRY_SOURCES)
    dv_src.prompt = "Entry source"
    ws.add_data_validation(dv_src)
    dv_src.add(f"{col_letter('Entry source')}2:{col_letter('Entry source')}{LAST_ROW}")

    # --- Reference sheet ----------------------------------------------------
    ref = wb.create_sheet("Reference")
    ref.column_dimensions["A"].width = 26
    ref.column_dimensions["B"].width = 90
    rows = [
        ("Dongfeng Experience — Member workbook", ""),
        ("", ""),
        ("Tab name", f'Keep the member tab named exactly "{TAB}". Set DFENG_SHEETS_TAB_NAME to match.'),
        ("Header row", "Row 1 must stay EXACTLY as generated (12 columns). The bot checks the header on "
                       "startup and refuses to run if it drifts. Do not rename, reorder, or insert columns."),
        ("Bot-owned (cols 1-8)", "Written by the bot via the service account. Treat as read-only: "
                                 + ", ".join(BOT_COLUMNS)),
        ("Admin-owned (cols 9-12)", "Edited by humans only; the bot never writes these: "
                                    + ", ".join(ADMIN_COLUMNS)),
        ("Key column", f'"{KEY_COLUMN}" — unique per member; the bot upserts on it.'),
        ("PII columns", "Optional phone / Optional plate — keep private under Singapore PDPA; never share in chat. "
                        "See docs/pdpa-policy.md."),
        ("Allowed Tag values", ", ".join(TAGS)),
        ("Allowed Entry source values", ", ".join(ENTRY_SOURCES)),
        ("Retention", "Keep a member's row until they leave the community or request deletion; then clear/redact "
                      "their row and mark 'Deletion requested'. See docs/pdpa-policy.md."),
        ("", ""),
        ("Upload to Google Drive", "1) Upload this .xlsx to Drive.  2) Open it, File > Save as Google Sheets "
                                   "(or right-click > Open with > Google Sheets).  3) Confirm the tab is named "
                                   f'"{TAB}".  4) Share the Google Sheet with the service-account email as Editor '
                                   "AND with each named admin's Google account.  5) Copy the sheet id from its URL "
                                   "(between /d/ and /edit) into DFENG_SHEETS_WORKBOOK_ID, and set "
                                   "DFENG_FEATURE_SHEETS=1.  Full steps: docs/google-sheets-setup.md."),
    ]
    for r, (a, b) in enumerate(rows, start=1):
        ca = ref.cell(row=r, column=1, value=a)
        cb = ref.cell(row=r, column=2, value=b)
        cb.alignment = Alignment(wrap_text=True, vertical="top")
        if r == 1:
            ca.font = Font(bold=True, size=13, color="1F3864")
        elif b:
            ca.font = Font(bold=True)

    wb.save(OUT)
    print(f"Wrote {OUT.relative_to(ROOT)}")
    print("Members header:", MEMBER_COLUMNS)


if __name__ == "__main__":
    build()
